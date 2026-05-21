"""Excel parser — each non-empty row in each sheet becomes a TypedContentItem.

Content is tab-joined cell values. ExcelLocator carries (sheet, cells) where
cells is in A1:B12 style — the first non-None column through the last.
"""

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from dks.locators import ExcelLocator
from dks.types import TypedContentItem


def parse_excel_file(path: Path) -> list[TypedContentItem]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    items: list[TypedContentItem] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = list(row)
            # Skip fully-empty rows (all None or all blank strings)
            if all(c is None or (isinstance(c, str) and c.strip() == "") for c in cells):
                continue
            # Trim trailing None
            while cells and cells[-1] is None:
                cells.pop()
            if not cells:
                continue
            last_col_letter = get_column_letter(len(cells))
            cell_range = f"A{row_idx}:{last_col_letter}{row_idx}"
            content = "\t".join("" if c is None else str(c) for c in cells)
            items.append(
                TypedContentItem(
                    content=content,
                    block_type="table",
                    locator=ExcelLocator(sheet=sheet_name, cells=cell_range),
                )
            )
    return items
