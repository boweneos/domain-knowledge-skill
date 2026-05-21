from pathlib import Path

from openpyxl import Workbook

from dks.locators import ExcelLocator
from dks.parsers.excel import parse_excel_file


def _make_xlsx(path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Mortality"
    sheet.append(["Age", "Rate"])
    sheet.append([20, 0.001])
    sheet.append([21, 0.0012])
    wb.create_sheet("Lapse")
    wb["Lapse"].append(["Policy", "Year1"])
    wb["Lapse"].append(["Term-10", 0.07])
    wb.save(path)


def test_parse_excel_yields_one_item_per_nonempty_row(tmp_path):
    src = tmp_path / "assumptions.xlsx"
    _make_xlsx(src)
    items = parse_excel_file(src)
    # Mortality: 3 rows (header + 2 data); Lapse: 2 rows (header + 1 data) → 5 total
    assert len(items) == 5
    sheets = {i.locator.sheet for i in items if isinstance(i.locator, ExcelLocator)}
    assert sheets == {"Mortality", "Lapse"}


def test_parse_excel_locator_carries_sheet_and_cells(tmp_path):
    src = tmp_path / "a.xlsx"
    _make_xlsx(src)
    items = parse_excel_file(src)
    first = items[0]
    assert isinstance(first.locator, ExcelLocator)
    assert first.locator.sheet == "Mortality"
    assert first.locator.cells == "A1:B1"
    assert first.block_type == "table"


def test_parse_excel_content_is_tab_joined(tmp_path):
    src = tmp_path / "a.xlsx"
    _make_xlsx(src)
    items = parse_excel_file(src)
    header = items[0]
    assert header.content == "Age\tRate"


def test_parse_excel_skips_fully_empty_rows(tmp_path):
    src = tmp_path / "gaps.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "S"
    sheet.append(["A"])
    sheet.append([None])
    sheet.append(["B"])
    wb.save(src)
    items = parse_excel_file(src)
    assert [i.content for i in items] == ["A", "B"]
